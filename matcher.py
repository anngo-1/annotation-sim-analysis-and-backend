import json
from openpyxl import load_workbook, Workbook
import sys

wb = Workbook()
ws = wb.active
ws.title = "img-nose-matching"
ws.cell(1, 1, "Face")
ws.cell(1, 2, "Nose")

if len(sys.argv) != 4:
    print(f"Usage: python {sys.argv[0]} <fairface_sheet> <bitmoji_sheet> <export_sheet>")
    sys.exit()

fairface = load_workbook(sys.argv[1]).active
bitmoji = load_workbook(sys.argv[2]).active
output = sys.argv[3]

y = 2
for annotation in fairface.iter_rows(min_row=1, max_col=3, max_row=5, values_only=True):
    tags1 = json.loads(annotation[2])
    matches = list()
    best = 0
    for row in bitmoji.iter_rows(min_row=1, max_col=3, max_row=51, values_only=True):
        tags2 = json.loads(row[1])
        score = 0
        for i in range(len(tags1)):
            if tags1[i] == tags2[i]:
                score += 1
        if score > best:
            matches.clear()
            matches.append(row[0][:-4])
            best = score
        elif score == best:
            matches.append(row[0][:-4])
    ws.cell(y, 1, annotation[1][:-4])
    ws.cell(y, 2, ', '.join(matches))
    print(annotation[1][:-4] + " : " + ', '.join(matches))
    y += 1
wb.save(output)
    
